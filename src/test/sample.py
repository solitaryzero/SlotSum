import json
import random
import os
import openpyxl
import shutil


if __name__ == '__main__':
    sloted_data_file = './data/slotsum/sloted/test.jsonl'
    with open(sloted_data_file, 'r', encoding='utf-8') as fin:
        all_data = []
        for i, line in enumerate(fin):
            js = json.loads(line)
            all_data.append((i, js))

    sample_num = 20
    sampled_data = random.sample(all_data, sample_num)

    prediction_base_path = './data/slotsum/predictions'
    prediction_paths = {
        'BART-raw': '%s/bart_raw/prediction.jsonl' %prediction_base_path,
        # 'BART-cnn': '%s/bart_cnn/prediction.jsonl' %prediction_base_path,
        'T5': '%s/t5-base/prediction.jsonl' %prediction_base_path,
        'Pegasus': '%s/pegasus/prediction.jsonl' %prediction_base_path,
        'NoisySumm': '%s/noisysumm/prediction.jsonl' %prediction_base_path,
        'SlotSum_predict': '%s/template_prediction_nocnn/filled_predictions_predict.jsonl' %prediction_base_path,
        'SlotSum_allpredict': '%s/template_prediction_nocnn/filled_predictions_allpredict.jsonl' %prediction_base_path,
        'SlotSum_discard': '%s/template_prediction_nocnn/filled_predictions_discard.jsonl' %prediction_base_path,
        # 'SlotSum_keep': '%s/template_prediction_nocnn/filled_predictions_keep.jsonl' %prediction_base_path,
    }

    out_path = './data/annotation'
    reference_path = './data/reference'

    # clean garbages
    if not(os.path.exists(out_path)):
        os.makedirs(out_path)
    for file in os.listdir(out_path):
        path = os.path.join(out_path, file)
        try:
            shutil.rmtree(path)
        except OSError:
            os.remove(path)

    if not(os.path.exists(reference_path)):
        os.makedirs(reference_path)
    for file in os.listdir(reference_path):
        path = os.path.join(reference_path, file)
        try:
            shutil.rmtree(path)
        except OSError:
            os.remove(path)

    for model in prediction_paths:
        all_prediction_file = prediction_paths[model]
        all_predictions = []
        with open(all_prediction_file, 'r', encoding='utf-8') as fin:
            for line in fin:
                all_predictions.append(json.loads(line))

        out_file_path = os.path.join(out_path, model)
        if not(os.path.exists(out_file_path)):
            os.makedirs(out_file_path)

        for index, data in sampled_data:
            backgrounds = {
                'title': data['title'],
                'golden': all_predictions[index]['text'],
                'prediction': all_predictions[index]['claim'],
                'replaced_slots': data['replaced_slots'],
            }
            
            questions = []
            q_num = min(len(backgrounds['replaced_slots']), 3)
            replaced_slots = random.sample(backgrounds['replaced_slots'], q_num)
            for key, value in replaced_slots:
                q_key = key.lower().replace('_', ' ')
                question = 'What is the %s mentioned in this text?' %q_key
                answer = value
                questions.append((question, answer))

            # Generate xlsx file for annotation
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.cell(1, 1, 'Background:')
            worksheet.cell(2, 1, backgrounds['prediction'])

            worksheet.cell(3, 1, 'Linguistic Scores:')
            worksheet.cell(4, 1, 'Completeness')
            worksheet.cell(4, 2, 'Fluency')
            worksheet.cell(4, 3, 'Succinctness')

            worksheet.cell(6, 1, 'Questions:')
            for j, q in enumerate(questions):
                worksheet.cell(7+j*2, 1, '%d. %s' %(j+1, q[0]))

            workbook.save(os.path.join(out_file_path, '%d.xlsx' %(index)))


    reference_file_path = reference_path
    if not(os.path.exists(reference_file_path)):
        os.makedirs(reference_file_path)

    for index, data in sampled_data:
        backgrounds = {
            'title': data['title'],
            'golden': all_predictions[index]['text'],
            'prediction': all_predictions[index]['claim'],
            'replaced_slots': data['replaced_slots'],
        }
            
        questions = []
        q_num = min(len(backgrounds['replaced_slots']), 3)
        replaced_slots = random.sample(backgrounds['replaced_slots'], q_num)
        for key, value in replaced_slots:
            q_key = key.lower().replace('_', ' ')
            question = 'What is the %s mentioned in this text?' %q_key
            answer = value
            questions.append((question, answer))
        # Generate xlsx file for reference
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.cell(1, 1, 'Background:')
        worksheet.cell(2, 1, backgrounds['golden'])
        worksheet.cell(4, 1, 'Questions:')
        for j, q in enumerate(questions):
            worksheet.cell(5+j*2, 1, '%d. %s' %(j+1, q[0]))
            # worksheet.cell(5+j*2+1, 1, q[1])

        workbook.save(os.path.join(reference_file_path, '%d.xlsx' %(index)))
